from openpyxl import Workbook, load_workbook 
wb = load_workbook('tests/test1.xlsx')
ws = wb.active
total = 0
max_row = ws.max_row
for i in range(2, max_row + 1):
    hours = ws['C'+str(i)].value
    rate = ws['B'+str(i)].value
    if isinstance(hours, int) and isinstance(rate, int):
        salary = float(hours) * float(rate)
        ws['D' + '4'].value = salary
        if salary > 3000:
            total += 1
print("The amount of people whose salary is highther than 3000, are:",total)
wb.save('result.xcxl')
wb.close()
